from tkinter import *
import tkinter.messagebox
import tkinter.ttk as ttk
from tkinter.filedialog import askopenfilename, asksaveasfilename
import threading
import os
from openpyxl import load_workbook
import xlrd
import xlwt
import _pickle

from DateUtils import DateEntry
from DateUtils import CalendarDialog
import datetime
from ParseCommon import Parser, curr_date, start_search_msg, cancel_search_msg, finish_search_msg
from ParseOrbita import ParserOrbita
from ParseZahav import ParserZahav
from ParseSnukRu import ParserSnukRu
from ParseIsraCom import ParserIsraCom
from ParseIsraVid import ParserIsraVid
from ParseDoskaCoil import ParserDoskaCoil
from ParseDoskiCoil import ParserDoskiCoil
from ParseDoskaIsraelInfoCo import ParserDoskaIsraelInfoCo
from ParseRabotaCoIl import ParserRabotaCoIl
from ParseBoardIinfo import ParserBoardIinfo
from ParseLemhira import ParserLemhira
from ParseSova import ParserSova
from ParseIlBoard import ParserIlBoard
from ParseRussianDoska import ParserRussianDoska
from ParseSoyuz import ParserSoyuz
from ParseRabotaIsrael import ParserRabotaIsrael
from ParseBoardIsra import ParserBoardIsra
from ParseBuySell import ParserBuySell


class ErrorLogger(object):
    def __init__(self, filename):
        self.terminal = sys.stderr
        self.log = open(filename, "a+")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)


class MainWindow(object):

    def __init__(self):
        self.running = False
        self.start_time = curr_date()
        self.timer = None
        self.controls = list()
        self.source_filename = None
        topframe = Frame(master)
        topframe.pack(side=TOP, fill=None)
        Label(topframe, text="Поиск телефонов начиная с").pack(side=LEFT, padx=2, pady=2)
        self.date_entry = DateEntry.DateEntry(topframe, default_date=datetime.datetime.today())
        self.date_entry.pack(side=LEFT, padx=2, pady=2)
        self.controls.append(self.date_entry.entry_1)
        self.controls.append(self.date_entry.entry_2)
        self.controls.append(self.date_entry.entry_3)
        btn_date = ttk.Button(topframe, text="..", command=self.btn_date_command, width=2)
        btn_date.pack(side=LEFT, padx=2, pady=2)
        self.controls.append(btn_date)

        excelframe = Frame(master)
        excelframe.pack(side=TOP, fill=X, expand=1)
        Label(excelframe, text="Excel-файл с существующими телефонами").pack(side=LEFT, padx=2, pady=2)
        self.excel_sv = StringVar()
        self.excel_entry = Entry(excelframe, textvariable=self.excel_sv)
        self.excel_entry.pack(side=LEFT, padx=2, pady=2, expand=1, fill=X)
        self.excel_entry.bind("<Key>", lambda e: "break")
        self.controls.append(self.excel_entry)
        btn_excel = ttk.Button(excelframe, text="..", command=self.btn_excel_command, width=2)
        btn_excel.pack(side=LEFT, padx=2, pady=2)
        self.controls.append(btn_excel)

        main_frame = Frame(master, relief=SOLID, bd=1)
        main_frame.pack(side=TOP, expand=1, fill=BOTH)
        Label(main_frame, text="Фильтр", bd=1, relief=SOLID, bg="silver").grid(row=0, column=0, padx=2, pady=2, sticky=N+S+E+W)
        Label(main_frame, text="Сайт", bd=1, relief=SOLID, bg="silver").grid(row=0, column=1, padx=2, pady=2, sticky=N+S+E+W)
        Label(main_frame, text="Страниц\nобработано", bd=1, relief=SOLID, bg="silver").grid(row=0, column=2, padx=2, pady=2, sticky=N+S+E+W)
        Label(main_frame, text="Телефонов\nнайдено", bd=1, relief=SOLID, bg="silver").grid(row=0, column=3, padx=2, pady=2, sticky=N+S+E+W)
        Label(main_frame, text="Ошибок", bd=1, relief=SOLID, bg="silver").grid(row=0, column=4, padx=2, pady=2, sticky=N+S+E+W)
        Label(main_frame, text="Статус", bd=1, relief=SOLID, bg="silver").grid(row=0, column=5, padx=2, pady=2, sticky=N+S+E+W)
        
        for i in range(len(parsers)):

            bv = BooleanVar()
            bv.set(i not in (5,6))
            cb = Checkbutton(main_frame, bd=1, relief=SOLID, bg="white", variable=bv)
            cb.grid(row=i+1, column=0, padx=2, pady=2, sticky=N+S+E+W)
            parsers[i]['checked'] = bv
            self.controls.append(cb)

            sv = StringVar()
            sv.set(parsers[i]['parser'].base_url)
            lb = Label(main_frame, textvariable=sv, bd=1, relief=SOLID, bg="white")
            lb.grid(row=i+1, column=1, padx=2, pady=2, sticky=N+S+E+W)

            sv = StringVar()
            sv.set('0')
            lb = Label(main_frame, textvariable=sv, bd=1, relief=SOLID, bg="white")
            lb.grid(row=i+1, column=2, padx=2, pady=2, sticky=N+S+E+W)
            parsers[i]['pages'] = sv
            parsers[i]['pages_lb'] = lb

            sv = StringVar()
            sv.set('0')
            lb = Label(main_frame, textvariable=sv, bd=1, relief=SOLID, bg="white")
            lb.grid(row=i+1, column=3, padx=2, pady=2, sticky=N+S+E+W)
            parsers[i]['phones'] = sv
            parsers[i]['phones_lb'] = lb

            sv = StringVar()
            sv.set('0')
            lb = Label(main_frame, textvariable=sv, bd=1, relief=SOLID, bg="white")
            lb.grid(row=i+1, column=4, padx=2, pady=2, sticky=N+S+E+W)
            parsers[i]['errors'] = sv
            parsers[i]['errors_lb'] = lb

            sv = StringVar()
            sv.set('')
            lb = Label(main_frame, textvariable=sv, bd=1, relief=SOLID, bg="white", width=15)
            lb.grid(row=i+1, column=5, padx=2, pady=2, sticky=N+S+E+W)
            parsers[i]['status'] = sv
            parsers[i]['status_lb'] = lb

        self.bv_all = BooleanVar()
        self.bv_all.set(True)
        self.cb_all = Checkbutton(main_frame, bd=1, relief=SOLID, bg="silver", activebackground="silver", command=self.cb_all_command, variable=self.bv_all)
        self.cb_all.grid(row=i+2, column=0, padx=2, pady=2, sticky=N+S+E+W)
        self.controls.append(self.cb_all)
        self.lb_pages = Label(main_frame, text="Итого:", bd=1, relief=SOLID, bg="silver")
        self.lb_pages.grid(row=i+2, column=1, padx=2, pady=2, sticky=N+S+E+W)
        self.lb_pages = Label(main_frame, text="0", bd=1, relief=SOLID, bg="silver")
        self.lb_pages.grid(row=i+2, column=2, padx=2, pady=2, sticky=N+S+E+W)
        self.lb_phones = Label(main_frame, text="0", bd=1, relief=SOLID, bg="silver")
        self.lb_phones.grid(row=i+2, column=3, padx=2, pady=2, sticky=N+S+E+W)
        self.lb_errors = Label(main_frame, text="0", bd=1, relief=SOLID, bg="silver")
        self.lb_errors.grid(row=i+2, column=4, padx=2, pady=2, sticky=N+S+E+W)
        self.lb_time = Label(main_frame, text="0:00:00", bd=1, relief=SOLID, bg="silver")
        self.lb_time.grid(row=i+2, column=5, padx=2, pady=2, sticky=N+S+E+W)

        bottomframe = Frame(master)
        bottomframe.pack(side=BOTTOM)
        self.btn_start = ttk.Button(bottomframe, text="Начать поиск", command=self.btn_start_command)
        self.btn_start.pack(side=LEFT, padx=2, pady=6)
        self.btn_save = ttk.Button(bottomframe, text="Сохранить результат", command=self.btn_save_command)
        self.btn_save.pack(side=LEFT, padx=2, pady=6)
        self.controls.append(self.btn_save)

    def cb_all_command(self):
        for p in parsers:
            p['checked'].set(self.bv_all.get())

    def btn_save_command(self):
        save_filename = asksaveasfilename(filetypes=(("Excel file", "*.xls"), ))
        if not save_filename:
            return
        filename, file_extension = os.path.splitext(save_filename)
        if file_extension == '':
            save_filename += '.xls'
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Найденные телефоны')

        phone_style = xlwt.XFStyle()
        phone_style.num_format_str = '0000000000'

        ws.col(0).width = 256 * 11

        line = 0
        for k in Parser.phones.keys():
            if k not in Parser.source_excel.keys():
                try:
                    v = Parser.phones[k]
                except EOFError:
                    v = ''
                ws.write(line, 0, int(k), phone_style)
                ws.write(line, 1, v)
                line += 1

        wb.save(save_filename)
        os.startfile(save_filename, 'open')

    def btn_date_command(self):
        
        """try:
            init_date = self.date_entry.get()
        except ValueError:
            init_date = None"""
        cd = CalendarDialog.CalendarDialog(master)
        result = cd.result
        self.date_entry.set_date(result)

    def load_xlsx(self):
        wb = load_workbook(self.source_filename, read_only=True)
        ws = wb.active
        for row in ws.rows:
            for cell in row:
                phone_str = str(cell.value).strip().zfill(10)
                Parser.source_excel[phone_str] = Parser.source_excel.get(phone_str, 0) + 1
                break

    def load_xls(self):
        rb = xlrd.open_workbook(self.source_filename)
        sheet = rb.sheet_by_index(0)
        for rownum in range(sheet.nrows):
            row = sheet.row_values(rownum)
            for cell in row:
                if isinstance(cell, float):
                    cell = int(cell)
                phone_str = str(cell).strip().zfill(10)
                Parser.source_excel[phone_str] = Parser.source_excel.get(phone_str, 0) + 1
                break

    def btn_excel_command(self):
        self.source_filename = askopenfilename(filetypes=(("Excel files", "*.xls;*.xlsx"), ))
        self.excel_sv.set(self.source_filename)
        filename, file_extension = os.path.splitext(self.source_filename)
        Parser.source_excel = dict()
        if file_extension == '.xlsx':
            self.load_xlsx()
        else:
            self.load_xls()

    def update_elapsed_time(self):
        elapsed_time = datetime.datetime.today() - self.start_time
        hours, seconds = divmod(elapsed_time.seconds, 3600)
        minutes, seconds = divmod(seconds, 60)
        self.lb_time.configure(text=str(hours)+':'+str(minutes).zfill(2)+':'+str(seconds).zfill(2))
        if self.running:
            self.timer = threading.Timer(1.0, self.update_elapsed_time)
            self.timer.start()
        else:
            self.timer = None

    def btn_start_command(self):
        if self.running:
            self.running = False
            self.btn_start.configure(state=DISABLED)
            for p in parsers:
                if p['parser'].running:
                    p['parser'].cancel()
        else:
            if len([p for p in parsers if p['checked'].get()]) == 0:
                tkinter.messagebox.showerror("Ошибка", "Нет выбранных сайтов")
                return
            try:
                Parser.limit_date_for_search = self.date_entry.get()
            except ValueError:
                tkinter.messagebox.showerror("Ошибка", "Укажите корректную дату")
                return
            has_restoring = False
            for p in parsers:
                if p['checked'].get():
                    for k in Parser.cfg.keys():
                        try:
                            v = Parser.cfg[k]
                        except (EOFError, _pickle.UnpicklingError, KeyError):
                            v = ['']
                            Parser.cfg[k] = v
                        if isinstance(v[0], int):
                            if k.startswith(p['parser'].base_url) and v[0] != 0:
                                has_restoring = True
                        else:
                            if k.startswith(p['parser'].base_url) and v[0] != '':
                                has_restoring = True
            if has_restoring:
                if not tkinter.messagebox.askyesno('Внимание',
                                                   ('Предыдущий сеанс поиска был прерван. Возобновить поиск? '
                                                    '(Yes - возобновить, No - начать сначала)')):
                    for p in parsers:
                        if p['checked'].get():
                            for k, v in Parser.cfg.items():
                                if isinstance(v[0], int):
                                    if k.startswith(p['parser'].base_url) and v[0] != 0:
                                        Parser.cfg[k] = [0, v[1]]
                                else:
                                    if k.startswith(p['parser'].base_url) and v[0] != '':
                                        Parser.cfg[k] = ['', v[1]]

            self.running = True
            self.btn_start.configure(text='Отмена')
            self.btn_save.configure(state=DISABLED)
            self.start_time = datetime.datetime.today()
            self.update_elapsed_time()
            for item in self.controls:
                item.configure(state=DISABLED)

            for p in parsers:
                if p['checked'].get():
                    p['parser'].start()


def update_parser(i):
    result = False
    p = parsers[i]
    if str(p['parser'].page_count) != p['pages'].get():
        p['page_count'] = p['parser'].page_count
        p['pages'].set(str(p['page_count']))
        mainwindow.lb_pages.configure(text=str(sum([p['page_count'] for p in parsers if 'page_count' in p.keys()])))
        p['pages_lb'].configure(bg='#AAFFAA')
    else:
        p['pages_lb'].configure(bg='#FFFFFF')
    if str(p['parser'].phone_count) != p['phones'].get():
        p['phone_count'] = p['parser'].phone_count
        p['phones'].set(str(p['phone_count']))
        mainwindow.lb_phones.configure(text=str(sum([p['phone_count'] for p in parsers if 'phone_count' in p.keys()])))
        p['phones_lb'].configure(bg='#AAFFAA')
    else:
        p['phones_lb'].configure(bg='#FFFFFF')
    if str(p['parser'].error_count) != p['errors'].get():
        p['error_count'] = p['parser'].error_count
        p['errors'].set(str(p['error_count']))
        mainwindow.lb_errors.configure(text=str(sum([p['error_count'] for p in parsers if 'error_count' in p.keys()])))
        p['errors_lb'].configure(bg='#FFAAAA')
    else:
        p['errors_lb'].configure(bg='#FFFFFF')
    if p['parser'].status != p['status'].get():
        p['status'].set(p['parser'].status)
        if p['parser'].status == '':
            p['status_lb'].configure(bg="white")
        elif p['parser'].status == start_search_msg:
            p['status_lb'].configure(bg="#F7FFBB")
        elif p['parser'].status == cancel_search_msg:
            p['status_lb'].configure(bg="#FFAAAA")
            result = True
        elif p['parser'].status == finish_search_msg:
            p['status_lb'].configure(bg="#AAFFAA")
            result = True
        if p['parser'].status in (cancel_search_msg, finish_search_msg):
            finished = True
            for item in parsers:
                if item['parser'].running:
                    finished = False
                    break
            if finished:
                mainwindow.btn_start.configure(text='Начать поиск', state=NORMAL)
                mainwindow.running = False
                for item in mainwindow.controls:
                    item.configure(state=NORMAL)
                if mainwindow.timer is not None:
                    mainwindow.timer.cancel()
                    mainwindow.timer = None
    return result


def event_update_parser(p):
    index = p.index
    if update_parser(index):
        parsers[index]['parser'] = p()


def add_parser(new_parser_class):
    new_parser = new_parser_class()
    new_parser_class.index = len(parsers)
    parsers.append({'parser': new_parser})
    master.bind('<<'+new_parser.base_url+'>>', lambda x: event_update_parser(new_parser_class))


def center(toplevel):
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth()
    h = toplevel.winfo_screenheight()
    size = tuple(int(_) for _ in toplevel.geometry().split('+')[0].split('x'))
    x = w/2 - size[0]/2
    y = h/2 - size[1]/2
    toplevel.geometry("%dx%d+%d+%d" % (size + (x, y)))

sys.stderr = ErrorLogger(Parser.work_dir + '/error.log')
Parser.init_shelves()
mainwindow = None
try:
    master = Tk()
    master.title("Поиск телефонов")

    parsers = list()

    add_parser(ParserOrbita)
    add_parser(ParserZahav)
    add_parser(ParserSnukRu)
    add_parser(ParserIsraCom)
    add_parser(ParserIsraVid)
    add_parser(ParserDoskaCoil)
    add_parser(ParserDoskiCoil)
    add_parser(ParserDoskaIsraelInfoCo)
    add_parser(ParserRabotaCoIl)
    add_parser(ParserBoardIinfo)
    add_parser(ParserLemhira)
    add_parser(ParserSova)
    add_parser(ParserIlBoard)
    add_parser(ParserRussianDoska)
    add_parser(ParserSoyuz)
    add_parser(ParserRabotaIsrael)
    add_parser(ParserBoardIsra)
    add_parser(ParserBuySell)

    Parser.root = master

    mainwindow = MainWindow()

    master.resizable(0, 0)
    center(master)
    mainloop()
finally:
    if mainwindow is not None:
        if mainwindow.timer is not None:
            mainwindow.timer.cancel()

    Parser.close_shelves()
